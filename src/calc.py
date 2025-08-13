import pandas as pd
import os
from openpyxl import load_workbook
import yfinance as yf

def adjust_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

def main():
    file_path = r'data/percent_changes.xlsx'
    sectors = {
        "Communication": ["GOOGL", "META", "VZ"],
        "Consumer Discretionary": ['AMZN', 'HD', 'MCD'],
        "Consumer Staples": ['PG', 'KO', 'PEP'],
        "Energy": ['XOM', 'CVX', 'COP'],
        "Financials": ['JPM', 'BAC', 'WFC'],
        "Healthcare": ['JNJ', 'PFE', 'ABT'],
        "Industrials": ['HON', 'RTX', 'UNP'],
        "Materials": ['LIN', 'APD', 'SHW'],
        "Real Estate": ["PLD", "SPG", "O"],
        "Technology": ["NVDA", "MSFT", "AAPL"],
        "Utilities": ["NEE", "DUK", "D"]
    }

    with pd.ExcelWriter(file_path, mode='w', engine="openpyxl") as writer:
        for sector, tickers in sectors.items():
            dfs = []
            for ticker in tickers:
                stock = yf.Ticker(ticker)
                historical_data = stock.history(start="2024-12-02", end="2025-02-17")
                df1 = historical_data[['Open', 'Close']].copy()
                df1 = df1.reset_index()
                df1['Date'] = df1['Date'].dt.tz_localize(None)
                df1['%Open'] = ((df1['Open'] - df1['Close'].shift(1)) / df1['Close'].shift(1)) * 100
                df1['%Close'] = ((df1['Close'] - df1['Close'].shift(1)) / df1['Close'].shift(1)) * 100
                df1 = df1[['Date', '%Open', '%Close']]
                df1.columns = ['Date', f"{ticker} %Open", f"{ticker} %Close"]
                dfs.append(df1.set_index('Date'))
            df = pd.concat(dfs, axis=1)
            df.to_excel(writer, sheet_name=sector)

    workbook = load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        adjust_column_widths(sheet)
    workbook.save(file_path)

if __name__ == '__main__':
    main()
